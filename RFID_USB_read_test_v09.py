"""
coding: utf-8
v03 以下URLを参考に、同時複数Tag読み込みの精度を上げるため、リストにすべて入れてから重複分を削除する方法に改善
https://note.nkmk.me/python-list-unique-duplicate/
v04 excelに登録しているリストから内容照会、確認日記入まで
v06 更にexcelに書き込める機能を追加
v09 読み込んだID番号がリストに存在していない時にエラーメッセージを発生する機能追加
"""

import time
import tkinter
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import messagebox
import openpyxl
import serial
from datetime import datetime

path1 = '//telatnas50/0CST2/CE4_DEPT/DX/00_個人フォルダ/榎/RFID/'
file1 = 'rfid_sample.xlsx'
wb = openpyxl.load_workbook(path1 + file1)
sheet_list = wb['list (2)']
tag_id_list = []
# print('先頭にもどった')
# tag_id_list_saved = []

def read():
    i = 0
    tag_id_list = []
    ser = serial.Serial("COM4", 115200)
    while i < 10:
        if not ser.is_open:
            break

        coming = ser.read(22)
        # print(coming)
        # print(type(coming))
        hex_coming = coming.hex()
        # print(hex_coming)
        # print(type(hex_coming))
        # print(int(hex_coming))
        tag_id = hex_coming[26:38]
        # print(tag_id)
        tag_id_list.append(tag_id)  # v03 読んだTagをすべてリストに入れる
        i += 1

    # print(tag_id_list)
    # print(tag_id_list[0])
    global tag_id_list_unique
    tag_id_list_unique = list(set(tag_id_list))  # ここでリスト内の重複分を削除
    # print(tag_id_list_unique)
    text.set("Retry")
    text2.set("IDが読まれました。\n別のIDを読む場合は近づけて\nRetryを押してください。")
    tag_id_list_string = "\n".join(tag_id_list_unique)
    text3.set(tag_id_list_string)
    text4.set('読み取ったIDリスト')
    global tag_id_list_saved
    tag_id_list_saved = tag_id_list_string
    # print('tag_id_list_saved', tag_id_list_saved)
    ser.close()
    time.sleep(0.1)


def restart():  # リストを初期化して再読み込み
    tag_id_list = []
    ser = serial.Serial("COM4", 115200)  # 前の読み込みデータが残る対策で、一度シリアルOpenしCloseする
    time.sleep(0.1)
    ser.close()
    read()

# duplication = tag_id in tag_id_list #v03でマスク

def search_cell_value_vertical(start_row, column, sheet, texts):  # 読んだタグのIDがあるエクセルの行番号(row)を返す
    cell_value_list = []
    search_text_cell = 0

    for i in range(start_row, sheet.max_row + 1):
        # print(i, sheet["A" + str(i)].value)
        cell_value = sheet.cell(row=i, column=column).value  # row行、column列
        # print('cell_value', cell_value)
        cell_value_list.append(cell_value)
        # print('search cell i = ',i)
        # print('cell_value', cell_value)
        # print(type(cell_value))  #cell_value=str
        # print('text', texts)
        # print("cell_value_list", cell_value_list)
        if cell_value == texts:
            search_text_cell = i
            # print('i = 見つけるまで繰り返した回数 = ', i)
        else:
            pass
    # print('cell_value_list in read_cell :', cell_value_list)
    return search_text_cell


def read_cell_value_horizontal(row, start_column, number_of_column, sheet):
    cell_value_list = []
    for i in range(start_column, start_column + number_of_column):
        cell_value = sheet.cell(row=row, column=i).value
        # print('cell_value',cell_value)
        cell_value_list.append(cell_value)
    # print('cell_value_list in read_cell :',cell_value_list)
    return cell_value_list


def display_excel_content():
    # print("tag_id_list_saved", tag_id_list_saved)
    global row_id_location
    row_id_location = search_cell_value_vertical(2, 1, sheet_list, tag_id_list_saved)
    # print('tag_id_list_saved', tag_id_list_saved)
    # print('row_id_location', row_id_location)

    if row_id_location == 0:
        messagebox.showerror("エラー", "リストに存在しません")
    else:
        rfid_cell_value_list = read_cell_value_horizontal(row_id_location, 1, 5, sheet_list)
        print(rfid_cell_value_list)
        rfid = rfid_cell_value_list[0]
        # print(rfid)
        part_name = rfid_cell_value_list[1]
        # print(part_name)
        part_number = rfid_cell_value_list[2]
        # print(part_number)
        part_location = rfid_cell_value_list[3]
        # print(part_location)
        part_check_date = rfid_cell_value_list[4]
        # print(part_check_date)

        text7.set(rfid_cell_value_list[0] + '\n' + rfid_cell_value_list[1] + '\n' + rfid_cell_value_list[2] + '\n' +
                  rfid_cell_value_list[3] + '\n' + str(rfid_cell_value_list[4]))



def status_ok():
    timestamp = datetime.today()
    timename = timestamp.strftime("%Y/%m/%d %H:%M:%S")
    # print("timename", timename)
    if row_id_location == 0:
        messagebox.showerror("エラー", "リストに存在しません")
    else:
        text9.set("done " + "\n" +timename)
        # print(sheet_list.cell(row=row_id_location,column=1).value)
        sheet_list.cell(row=row_id_location,column=6).number_format = "YY/MM/DD"
        sheet_list.cell(row=row_id_location,column=6).value = timestamp
        # print("date:", sheet_list.cell(row=row_id_location,column=6).value)
        wb.save(path1 + file1)  # excelファイルの保存実施

def signatures():
    if row_id_location == 0:
        messagebox.showerror("エラー", "リストに存在しません")
    else:
        values_6 = entry6.get()
        # print("values_6", values_6)
        # sheet_list.cell(row=row_id_location, column=7).number_format = openpyxl.styles.numbers.FORMAT_TEXT
        sheet_list.cell(row=row_id_location, column=7).value = values_6
        wb.save(path1 + file1)  # excelファイルの保存実施
        messagebox.showinfo("完了", "シートに保存しました。\n別IDを読み込む場合は\nRetryボタンから始めてください。")

def quit():
    root.destroy()


# 参照URL　https://office54.net/python/tkinter/python-tkinter-button
# rootメインウィンドウの設定
root = tk.Tk()
root.title('確認')
root.geometry("500x800")
# メインフレームの作成と設置
frame = ttk.Frame(root)
frame.pack(fill=tk.BOTH, padx=100, pady=60)
# Treeview生成
tree = ttk.Treeview(root)
# StringVarのインスタンスを格納する変数textの設定
text = tk.StringVar(frame)
text.set("Start")
text2 = tk.StringVar(frame)
text2.set("")
text3 = tk.StringVar(frame)
text3.set("")
text4 = tk.StringVar(frame)
text4.set("")
text5 = tk.StringVar(frame)
text5.set("終了")
text6 = tk.StringVar(frame)
text6.set("情報照会")
text7 = tk.StringVar(frame)
text7.set("")
text8 = tk.StringVar(frame)
text8.set("check")
text9 = tk.StringVar(frame)
text9.set("")
# text10 = tk.StringVar(frame)
# text10.set("name")
text10 = tk.StringVar(frame)
text10.set("")
text11 = tk.StringVar(frame)
text11.set("")

# フレームの作成
frame1 = tk.LabelFrame(root, text="RFIDの読み取り情報", foreground="green")
frame2 = tk.LabelFrame(root, text="excel表との情報照会", foreground="green")
frame3 = tk.LabelFrame(root, text="記名", foreground="green")

# 各種ウィジェットの作成
button = tk.Button(frame, textvariable=text, command=restart, font=("MSゴシック", "12"))
label = tk.Label(frame, textvariable=text2, font=("MSゴシック", "12"))
label2 = tk.Label(frame, textvariable=text3, font=("MSゴシック", "14"),bg="white")
label3 = tk.Label(frame, textvariable=text4, font=("MSゴシック", "14"),bg="white")
button2 = tk.Button(frame, textvariable=text6, command=display_excel_content, font=("MSゴシック", "12"))
label4 = tk.Label(frame, textvariable=text7, font=("MSゴシック", "14"),bg="white")
button3 = tk.Button(frame, textvariable=text5, command=quit, font=("MSゴシック", "12"))
button4 = tk.Button(frame, textvariable=text8, command=status_ok, font=("MSゴシック", "12"))
label5 = tk.Label(frame, textvariable=text9, font=("MSゴシック", "14"),bg="white")
label8 = tk.Label(frame, textvariable=text10, font=("MSゴシック", "14"))
label6 = ttk.Label(frame, text="↓名前を記入し実行ボタン押してください", font=("MSゴシック", "12"))
entry6 = ttk.Entry(frame)
# style = ttk.Style()
# style.configure("office.TButton", font=5)
button6 = ttk.Button(frame, text="実行", command=signatures)
label7 = tk.Label(frame, textvariable=text11, font=("MSゴシック", "14"))

# 各種ウィジェットの配置
label.grid(column=0,sticky=tk.W+tk.E)
button.grid(column=0,sticky=tk.W+tk.E)
label3.grid(column=0,sticky=tk.W+tk.E)
label2.grid(column=0,sticky=tk.W+tk.E)
button2.grid(column=0,sticky=tk.W+tk.E)
label4.grid(column=0,sticky=tk.W+tk.E)
button4.grid(column=0,sticky=tk.W+tk.E)
label5.grid(column=0,sticky=tk.W+tk.E)
label8.grid(column=0,sticky=tk.W+tk.E)
label6.grid(row=9, column=0,sticky=tk.W+tk.E)
entry6.grid(row=10, column=0,sticky=tk.W+tk.E)
# entry6.grid(column=0,sticky=tk.W+tk.E)
# entry6.insert(tkinter.END,u"名前を記入し実行ボタン押してください")
# button6.grid(row=9, column=2)
button6.grid(column=0,sticky=tk.W+tk.E)
label7.grid(column=0,sticky=tk.W+tk.E)
button3.grid(column=0,sticky=tk.W+tk.E)


root.mainloop()